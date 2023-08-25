import win32com.client
import time
import pyautogui
import pandas as pd
import pyperclip
import openpyxl
from openpyxl import load_workbook

def process_data(df, linha_atual):
    df = df.dropna(subset=['Nota'])

    df.loc[:, 'Nota'] = df['Nota'].astype(str).str.rstrip('.0')

    next_row = df['Texto Longo Nota'].last_valid_index() + linha_atual + 1 if not df['Texto Longo Nota'].isnull().all() else linha_atual + 1

    wb = load_workbook('Backlog pendente.xlsx')
    ws = wb['Fluid']
    col_name = None
    for cell in ws[1]:
        if cell.value == 'Texto Longo Nota':
            col_name = cell.column_letter
            break

    if col_name is None:
        print("Coluna 'Texto Longo Nota' não encontrada.")
        return

    nota_linha_atual = df.loc[linha_atual - 1, 'Nota']

    pyperclip.copy(nota_linha_atual)

    for window in pyautogui.getWindowsWithTitle('SAP Easy access'):
        if "SAP" in window.title:
            pyautogui.hotkey('winleft', 'd')
            window.maximize()
            sapguiauto = win32com.client.GetObject('SAPGUI')
            application = sapguiauto.GetScriptingEngine
            connection = application.Children(0)
            session = connection.Children(0)
            session.findById('wnd[0]').maximize
            session.findById('wnd[0]/tbar[0]/okcd').text = 'IW32'
            time.sleep(0.2)
            pyautogui.press('enter')
            time.sleep(1)
            time.sleep(0.5)
            pyautogui.hotkey('ctrl', 'v')
            pyautogui.press('enter')
            time.sleep(1)
            pyautogui.click(x = 103, y = 200)
            time.sleep(0.4)
            pyautogui.press('tab')
            time.sleep(0.2)
            pyautogui.press('tab')
            pyautogui.press('enter')
            time.sleep(3.6)
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(0.1)
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(0.4)
            pyautogui.press('esc')
            time.sleep(0.8)
            pyautogui.press('esc')
            time.sleep(0.8)
            pyautogui.press('enter')
        for _ in range(7):
            pyautogui.press('esc')
            wb = load_workbook('Backlog pendente.xlsx')
            ws = wb['Fluid']

            cell = ws[f"{col_name}{next_row}"]

            cell.value = pyperclip.paste()

            wb.save('Backlog pendente.xlsx')
            wb.close()

linha_atual = 1

df = pd.read_excel('Backlog pendente.xlsx', sheet_name='Fluid')
total_linhas = len(df)

while linha_atual < total_linhas:
    process_data(df, linha_atual)
    linha_atual += 1



